import requests
from bs4 import BeautifulSoup
import openpyxl
import json
import re
import time
import argparse
import sys
import os
from datetime import datetime
# --- ПУТЬ К ТЕКУЩЕМУ СКРИПТУ ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# --- ПАПКА ДЛЯ ФАЙЛОВ ---
EXCEL_DIR = os.path.join(BASE_DIR, "excel_category")

# создаём папку, если она не существует
os.makedirs(EXCEL_DIR, exist_ok=True)

BASE_URL = "https://lalafo.kg"
START_URL = "https://lalafo.kg/kyrgyzstan"
HEADERS = { "User-Agent": "Mozilla/5.0" }
sys.stdout.reconfigure(encoding='utf-8')
# Функция для получения URL по номеру категории
def get_category_urls(category_number):
    full_url = categories.get(category_number)
    if full_url:
        short_url = "/" + full_url.split("/")[-1]
        return full_url, short_url
    else:
        return None, None
        
def remove_duplicates_excel(input_file, url_column=1):
    """
    Удаляет дубли в Excel по URL (по умолчанию первый столбец) и перезаписывает исходный файл.

    :param input_file: путь к исходному Excel файлу
    :param url_column: номер столбца для проверки дубликатов (1 = A)
    """
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    seen = set()
    rows_to_keep = []

    # Пробегаем все строки (кроме заголовка)
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            rows_to_keep.append(row)  # сохраняем заголовок
            continue
        url = row[url_column - 1]  # столбец URL
        if url not in seen:
            seen.add(url)
            rows_to_keep.append(row)

    # Очищаем лист и записываем только уникальные строки
    sheet.delete_rows(1, sheet.max_row)
    for row in rows_to_keep:
        sheet.append(row)

    wb.save(input_file)
    print(f"✅ Дубликаты удалены. Файл обновлён: {input_file}")


# ----------- ВЫТАСКИВАЕМ ВСЕ КАТЕГОРИИ -----------
def get_all_categories():
    print("📁 Собираем список категорий...")
    r = requests.get(START_URL, headers=HEADERS)
    soup = BeautifulSoup(r.text, "html.parser")
    categories = set()
    for a in soup.select("a[href^='/kyrgyzstan/']"):
        href = a.get("href")
        if "/ads/" in href:
            continue
        full = BASE_URL + href
        categories.add(full)
    print(f"📌 Найдено категорий: {len(categories)}")
    return list(categories)

# ----------- ВЫТАСКИВАЕМ ССЫЛКИ НА ОБЪЯВЛЕНИЯ -----------
def parse_ads_from_page(url):
    r = requests.get(url, headers=HEADERS)
    if r.status_code != 200:
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    ads = set()
    for a in soup.select("a[href*='/ads/']"):
        href = a.get("href")
        if href.startswith("/"):
            ads.add(BASE_URL + href)
    return list(ads)

def find_text_detailed(url):
    """
    Расширенный поиск текста с диагностикой
    """
    try:
        #print(f"🔍 Поиск текста на странице: {url}")
        
        # Получаем HTML
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        response = requests.get(url, headers=headers, timeout=10)
        response.encoding = 'utf-8'  # Принудительно устанавливаем кодировку
        
        #print(f"📊 Статус код: {response.status_code}")
        #print(f"📏 Размер страницы: {len(response.text)} символов")
        
        # Проверяем кодировку
        #print(f"🔤 Кодировка: {response.encoding}")
        
        # Парсим HTML
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Удаляем скрипты и стили для чистого текста
        for script in soup(["script", "style"]):
            script.decompose()
        
        # Получаем чистый текст
        page_text = soup.get_text()
        page_text = ' '.join(page_text.split())  # Убираем лишние пробелы
        
        #print(f"📝 Текст страницы (первые 500 символов):")
        #print(page_text[:500])
        #print("-" * 50)
        
        # Несколько способов поиска
        search_phrases = [
            'Вас может заинтересовать',
            'Вас может заинтересовать'.lower(),
            'вас может заинтересовать',  # в нижнем регистре
        ]
        
        found = False
        for phrase in search_phrases:
            if phrase in page_text:
                #print(f'✅ Текст "{phrase}" найден на странице!')
                found = True
                # Найдем контекст
                index = page_text.find(phrase)
                context = page_text[max(0, index-50):index+100]
                print(f"📋 Контекст: ...{context}...")
                break
            elif phrase in response.text:
                #print(f'⚠️ Текст "{phrase}" найден в HTML, но возможно в тегах/атрибутах')
                found = True
                break
        
        if not found:
            #print('❌ Текст "Вас может заинтересовать" не найден')
            
            # Поиск похожих фраз
            similar_patterns = [
                r'вас[\s]+может[\s]+заинтересовать',
                r'Вас[\s]+может[\s]+заинтересовать',
                r'заинтересовать',
            ]
            
            for pattern in similar_patterns:
                matches = re.findall(pattern, page_text, re.IGNORECASE)
                if matches:
                    print(f'🔍 Найдены похожие фразы: {matches}')
        
        return found
        
    except Exception as e:
        print(f'❌ Ошибка: {e}')
        return False

    
# ----------- ВЫТАСКИВАЕМ ДАННЫЕ ОБЪЯВЛЕНИЯ -----------
def extract_user_data(url):
    try:
        r = requests.get(url, headers=HEADERS)
        soup = BeautifulSoup(r.text, "html.parser")
        script_tag = soup.find("script", id="__NEXT_DATA__")
        if not script_tag:
            return None, None, None, None, None, None
        data = json.loads(script_tag.string)
        phone = find_deep(data, ["mobile", "phone", "telephone", "contact"])
        username = find_deep(data, ["username", "user_name"])
        city = find_deep(data, ["city", "City"])
        createdTime = find_deep(data, ["created_time"],allow_numbers=True)
        updatedTime = find_deep(data, ["updated_time"],allow_numbers=True)
        createdTimeStr = timestamp_to_date(createdTime)
        updatedTimeStr = timestamp_to_date(updatedTime)
        title = find_deep(data, ["title"])
        #print(f'🔍 createdTime =================================== {createdTimeStr}')
        if phone:
            phone = re.sub(r"[^\d+]", "", phone)
        return phone, username, city , createdTimeStr, updatedTimeStr, title
    except:
        return None, None, None, None, None, None


# ----------- РЕКУРСИВНЫЙ ПОИСК В JSON -----------
def find_deep(obj, keys, allow_numbers=False):
    if isinstance(obj, dict):
        for k in keys:
            if k in obj:
                value = obj[k]
                # Разрешаем числа или строки
                if (isinstance(value, str) and value.strip()) or (allow_numbers and isinstance(value, (int, float))):
                    return value
        for v in obj.values():
            result = find_deep(v, keys, allow_numbers)
            if result:
                return result
    elif isinstance(obj, list):
        for item in obj:
            result = find_deep(item, keys, allow_numbers)
            if result:
                return result
    return None

def timestamp_to_date(timestamp):
    """Конвертирует Unix timestamp в строку формата DD.MM.YY"""
    if timestamp:
        try:
            # Проверяем, в секундах или миллисекундах timestamp
            if timestamp > 10000000000:  # Если число очень большое - это миллисекунды
                timestamp = timestamp / 1000
            
            # Конвертируем timestamp в datetime объект
            dt = datetime.fromtimestamp(timestamp)
            # Форматируем в нужный формат
            return dt.strftime("%d.%m.%y")
        except Exception as e:
            print(f"❌ Ошибка конвертации timestamp {timestamp}: {e}")
            return None
    return None
    
# ----------- СОЗДАЁМ EXCEL -----------
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(["URL", "Название", "Пользователь", "Телефон", "Локация","Дата создания","Дата обновления" "Категория"])

# ----------- ЗАПУСК ПАРСИНГА С ЛОГИКОЙ СЧЁТЧИКА -----------
#categories = get_all_categories()

category = ""    
if len(sys.argv) > 1:
    category = sys.argv[1]
    print("Первый аргумент:", category)
    

cat_url = category
if cat_url:
    short_url = "/" + cat_url.split("/")[-1]
    print("Полный URL:", cat_url)
    print("Короткий URL:", short_url)
else:
    print("Категория с таким номером не найдена.")
    
page = 1
empty_pages = 0

# Убираем слэш в начале, чтобы имя файла было корректным
short_url_clean = short_url.lstrip("/")

# Формируем имя файла
filename = f"lalafo_ads_{short_url_clean}.xlsx"
exelFileName = f"lalafo_ads_{short_url_clean}.xlsx"
while True:
    page_url = f"{cat_url}?page={page}"
    print(f"  🔍 Страница {page}")
    isFinishTextFound = find_text_detailed(page_url)
    ads = parse_ads_from_page(page_url)
    if not ads:
        empty_pages += 1
        print(f"   🚫 Пустая страница {page}")
        if empty_pages >= 2:  # 2 пустые подряд → стоп
            break
    else:
        empty_pages = 0
        print(f"   ➡ Найдено объявлений: {len(ads)}")
        for ad in ads:
            print(f"     → Обрабатываем: {ad}")
            phone, user, city, createdTime, updatedTime, title = extract_user_data(ad)
            sheet.append([ad, title, user, phone, city, createdTime, updatedTime, cat_url])
            time.sleep(0.3)

    exelFileName = os.path.join(EXCEL_DIR, filename)
    wb.save(exelFileName)
    print(f"💾 Сохранено после страницы {page}")
    if page % 5 == 0:
        remove_duplicates_excel(exelFileName)
    page += 1
    
    if isFinishTextFound:
        break;
        
remove_duplicates_excel(exelFileName)
print(f"🎉 Готово! Категория спарсена и сохранена в {short_url_clean}")
